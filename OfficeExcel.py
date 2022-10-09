import iso8601
import sys
import os
import time
import datetime
from openpyxl import workbook as wb
from openpyxl import load_workbook as load_wb
from openpyxl.styles import Font, colors, Alignment

print(sys.modules['time'])


class OfficeExcel:
    def __init__(self, filepath: str, title_data=None):
        # self.filepath = ".//data//OfficeExcel.xlsx"

        self.filepath = filepath
        self.wb = None
        if os.path.exists(self.filepath):  # 检查文件是否存在，存在就读取，不存在就创建
            self.wb = load_wb(self.filepath)
        else:  # 初始化表格
            wb_temp = wb.Workbook()
            ws_temp = wb_temp.active
            # 创建表头
            # title_data = {"A1": "班级",
            #               "B1": "姓名",
            #               "C1": "学号",
            #               "D1": "联系方式",
            #               "E1": "学院/书院",
            #               "F1": "组织/部门",
            #               "G1": "借用物资",
            #               "H1": "借用日期",
            #               "I1": "归还日期",
            #               "J1": "是否已借出",
            #               "K1": "是否归还",
            #               "L1": "备注"}

            data_font = Font(name="等线", size=14, color=colors.BLACK, bold=True)
            data_align = Alignment(horizontal="center", vertical="center")
            for title in title_data.keys():
                ws_temp[title] = title_data[title]
                ws_temp[title].font = data_font
                ws_temp[title].alignment = data_align
            wb_temp.close()
            wb_temp.save(self.filepath)
            self.wb = load_wb(self.filepath)

        self.column_letter = []
        for letter in range(ord('A'), ord('Z')):
            self.column_letter.append(chr(letter))

        self.sheet = list()
        for sheet_name in self.wb.sheetnames:  # 获取全部表名
            self.sheet.append(sheet_name)

        # self.data = open(".\\data\\data.office", "a")
        self.ws = self.wb.active
        self.font = Font(name="等线", size=11, color=colors.BLACK)
        self.align = Alignment(horizontal="center", vertical="center")

    def __del__(self):
        # self.data.close()
        self.wb.close()
        pass

    def get_allsheel(self):
        return self.sheet

    def sheet_create(self, sheetname: str):
        self.wb.create_sheet(sheetname)
        self.sheet.append(sheetname)
        pass

    def sheet_choose(self, sheetname=None):
        if sheetname != None:
            self.ws = self.wb.get_sheet_by_name(sheetname)
            return True
        pass

    def cell_write(self, data: str, col: int, row: int):
        self.ws[(self.column_letter[col + 1] + str(row))] = data

    def data_write(self, data_row: list, now_row=0, isadd=True):
        if isadd:
            date_now = datetime.datetime.now()
            data_row.append(date_now.__format__("%Y%m%d-%H%M%S"))
            pass
        if now_row == 0:
            now_row = str(self.ws.max_row + 1)
        for i in range(0, data_row.__len__()):
            self.ws[self.column_letter[i] + now_row] = data_row[i]
            self.ws[self.column_letter[i] + now_row].font = self.font
            self.ws[self.column_letter[i] + now_row].alignment = self.align
        self.wb.save(self.filepath)
        pass

    def datas_write(self, data_2index: list):
        for data_row in data_2index:
            self.data_write(data_row)

    def data_read(self, min_row=2, max_row=0):
        ret = []
        if max_row == 0:
            max_row = self.ws.max_row
        for row in self.ws.iter_rows(min_row=min_row, max_row=max_row):
            row_temp = []
            for cell in row:
                if cell.value == None:
                    row_temp.append("")
                else:
                    row_temp.append(cell.value)
            ret.append(row_temp)
        return ret

    def data_find(self, column_letter: str, data, getpos=False):
        find_row = []
        find_index = []

        colnum = ord(column_letter) - ord('A') + 1
        coldata = self.ws.iter_cols(min_row=1, max_row=self.ws.max_row, min_col=colnum, max_col=colnum)
        coldata_once = list(coldata)[0]

        i = 0
        for celldata in coldata_once:
            try:
                #因为表格中不同数据类型不一，统一转化成str类型方便操作
                if str(celldata.value).find(str(data)) != -1:
                    find_index.append(i)
            except TypeError:
                try:
                    date = data.split("-")
                    data_datetime = \
                        datetime.date(year=int(date[0]), month=int(date[1]), day=int(data[2]))
                    if celldata.value.date().__eq__(data_datetime):
                        find_index.append(i)
                        pass
                except:
                    continue
            i += 1
        print(find_index)
        j = 0
        for row in self.ws.rows:
            if j in find_index:
                row_temp = []
                for cell in row:
                    row_temp.append(cell.value)
                find_row.append(row_temp)
            j += 1

        if getpos:
            ret_list = [find_row, i + 1]
            return ret_list
        if find_row.__len__() != 0:
            return find_row
        return []

    # if frist > second return True else return False
    def date_compare(self, frist_value: datetime, second_value: datetime):

        if frist_value.__gt__(second_value):
            return True
        else:
            return False
        # elif frist_value.__lt__(second_value):
        #     return False
        # return False

    def data_find_bydatetime(self, column_letter: str, data, start, end=None):
        if end is None:
            now_datetime = datetime.datetime.now() + datetime.timedelta(days=30)
            end = now_datetime

        ret = []
        data_find = self.data_find(column_letter=column_letter, data=data)
        print(data_find)
        for row in data_find:
            if (self.date_compare(row[7], start).__and__
                (self.date_compare(end, row[8]))):
                ret.append(row)

        return ret

        pass
